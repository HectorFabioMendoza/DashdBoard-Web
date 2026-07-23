# =============================================================================
# SCRIPT AUTOMATICO - Actualización del Tablero de Inventario Carapacho desde DBF
# Uso: python actualizar_inventario_carapacho.py
# Programar en el Programador de Tareas de Windows para correr 4 veces al día
# =============================================================================

import os
import sys
import logging
import shutil
import json
from datetime import datetime, date

# Intentar importar librerías requeridas o instalarlas automáticamente
try:
    from dbfread import DBF
except ImportError:
    print("Instalando librería dbfread...")
    os.system('pip install dbfread')
    from dbfread import DBF

try:
    import openpyxl
except ImportError:
    print("Instalando librería openpyxl...")
    os.system('pip install openpyxl')
    import openpyxl

# =============================================================================
# CONFIGURACION DE RUTAS
# =============================================================================
# Ruta principal en el servidor de Carapacho SM
RUTA_DATOS = r'F:\DataX\datos'
PROYECTO_DIR = os.path.dirname(os.path.abspath(__file__))

# Rutas de fallback para pruebas o red SMB desde Bodega
if not os.path.exists(RUTA_DATOS):
    fallbacks = [
        r'\\100.108.16.126\DataX\datos',
        r'E:\DataX_NUEVO\datos',
        r'D:\4 Hector Fabio\Distribuidora JR\Base de datos'
    ]
    for fb in fallbacks:
        if os.path.exists(fb):
            RUTA_DATOS = fb
            break

# Archivos DBF de origen
DBF_ITEMS   = os.path.join(RUTA_DATOS, 'initem.dbf')   # Catálogo de productos
DBF_MOVTOS  = os.path.join(RUTA_DATOS, 'inmvto.dbf')   # Ventas y movimientos
DBF_SALDOS  = os.path.join(RUTA_DATOS, 'insaldo.dbf')  # Saldos de inventario
DBF_STOCK   = os.path.join(RUTA_DATOS, 'instock.dbf')  # Fallback de stock por bodega

# Carpeta de salida (public para Vite dev/build)
PUBLIC_DIR = os.path.join(PROYECTO_DIR, 'public')
os.makedirs(PUBLIC_DIR, exist_ok=True)

FILE_INVENTARIO = os.path.join(PUBLIC_DIR, 'Inventario.xlsx')
FILE_VENTAS     = os.path.join(PUBLIC_DIR, 'Ventas por linea.xlsx')
FILE_LAST_UPDATE = os.path.join(PUBLIC_DIR, 'last_update.json')

# Rango de fechas dinámico para velocidad de ventas (últimos 12 meses)
FECHA_MIN = date(2025, 1, 1)
FECHA_MAX = date.today()

# =============================================================================
# INICIALIZACION DE LOGS
# =============================================================================
LOG_FILE = os.path.join(PROYECTO_DIR, 'log_inventario_carapacho.txt')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger()

# Parser seguro para manejar bytes nulos (\x00) en campos numéricos FoxPro
from dbfread.field_parser import FieldParser

class SafeFieldParser(FieldParser):
    def parseN(self, field, data):
        try:
            data_clean = data.replace(b'\x00', b'').strip()
            if not data_clean:
                return 0
            return super().parseN(field, data_clean)
        except ValueError:
            return 0

    def parseF(self, field, data):
        try:
            data_clean = data.replace(b'\x00', b'').strip()
            if not data_clean:
                return 0.0
            return super().parseF(field, data_clean)
        except ValueError:
            return 0.0

def leer_dbf(ruta):
    return DBF(ruta, load=False, encoding='latin-1', ignore_missing_memofile=True, parserclass=SafeFieldParser)

def safe_float(val, default=0.0):
    try:
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).strip()
        if not val_str:
            return default
        val_str = val_str.replace(',', '.')
        return float(val_str)
    except (ValueError, TypeError):
        return default

def obtener_serial_fecha(d):
    if d is None:
        return 0
    if isinstance(d, datetime):
        d = d.date()
    elif not isinstance(d, date):
        return 0
    fecha_base = date(1899, 12, 30)
    return (d - fecha_base).days

# Helper para buscar campos de forma insensible a mayúsculas
def get_field_val(row, keys_to_try, default=''):
    row_keys = {k.upper(): k for k in row.keys()}
    for k in keys_to_try:
        ku = k.upper()
        if ku in row_keys:
            val = row[row_keys[ku]]
            if val is not None:
                return val
    return default

# =============================================================================
# EJECUCION PRINCIPAL
# =============================================================================
def main():
    log.info("=================================================================")
    log.info(" INICIANDO ACTUALIZACION DEL TABLERO DE INVENTARIO CARAPACHO SM ")
    log.info(f" Ruta de datos DataX: {RUTA_DATOS}")
    log.info("=================================================================")

    if not os.path.exists(DBF_ITEMS):
        log.error(f"No se encontró el catálogo de ítems en {DBF_ITEMS}. Abortando.")
        sys.exit(1)

    # 0. Cargar Maestro de Grupos de Productos (Grupos GA01..GA22 de Siesa / DataX)
    log.info("Cargando maestro de grupos de productos (Grupos GA)...")
    grupos_ga_desc = {
        'GA01': 'PASTILLERÍA Y SALUD',
        'GA02': 'ARTÍCULOS DE BEBÉ',
        'GA03': 'ÚTILES ESCOLARES',
        'GA04': 'DESODORANTES',
        'GA05': 'JABONES DE BAÑO',
        'GA06': 'SHAMPOO Y ACONDICIONADOR',
        'GA07': 'MÁQUINAS DE AFEITAR',
        'GA08': 'CREMAS Y CEPILLOS DENTALES',
        'GA09': 'TOALLAS HIGIÉNICAS',
        'GA10': 'ARTÍCULOS DE ASEO',
        'GA11': 'DESECHABLES',
        'GA12': 'LICORES Y BEBIDAS',
        'GA13': 'DULCERÍA Y COMESTIBLES',
        'GA14': 'ARTÍCULOS DE TOCADOR',
        'GA15': 'ARTÍCULOS PARA ZAPATERÍA',
        'GA16': 'ARTÍCULOS DE MODISTERÍA',
        'GA17': 'ARTÍCULOS DE PVC / PLÁSTICOS',
        'GA18': 'ARTÍCULOS ELÉCTRICOS',
        'GA19': 'VARIOS Y DIVERSOS',
        'GA20': 'ARTÍCULOS DE CABELLO',
        'GA21': 'VITRINAS',
        'GA22': 'MOLDES PARA LLAVES'
    }

    clean_grupo_map = {
        'PASTILLERIA': 'PASTILLERÍA Y SALUD',
        'ARTICULOS DE BEBE': 'ARTÍCULOS DE BEBÉ',
        'UTILES ESCOLARES': 'ÚTILES ESCOLARES',
        'DESODORANTES': 'DESODORANTES',
        'JABONES DE BAÑO': 'JABONES DE BAÑO',
        'JABONES DE BA\ufffdO': 'JABONES DE BAÑO',
        'SHAMPOO': 'SHAMPOO Y ACONDICIONADOR',
        'MAQUINAS DE AFEITAR': 'MÁQUINAS DE AFEITAR',
        'CREMAS Y CEPILLOS D': 'CREMAS Y CEPILLOS DENTALES',
        'TOALLAS HIGIENICAS': 'TOALLAS HIGIÉNICAS',
        'ARTICULOS DE ASEO': 'ARTÍCULOS DE ASEO',
        'DESECHABLES': 'DESECHABLES',
        'LICORES': 'LICORES Y BEBIDAS',
        'DULCERIA-COMESTIBLE': 'DULCERÍA Y COMESTIBLES',
        'ARTICULOS TOCADOR': 'ARTÍCULOS DE TOCADOR',
        'ARTICULOS PARA ZAPA': 'ARTÍCULOS PARA ZAPATERÍA',
        'ARTICULOS MODISTERI': 'ARTÍCULOS DE MODISTERÍA',
        'ARTICULOS DE PVC': 'ARTÍCULOS DE PVC / PLÁSTICOS',
        'ARTICULOS ELECTRICO': 'ARTÍCULOS ELÉCTRICOS',
        'VARIOS': 'VARIOS Y DIVERSOS',
        'ARTICULOS DE CABELL': 'ARTÍCULOS DE CABELLO',
        'VITRINAS': 'VITRINAS',
        'MOLDES PARA LLAVES': 'MOLDES PARA LLAVES',
        '0201': 'PASTILLERÍA Y SALUD',
        '0202': 'ARTÍCULOS DE BEBÉ',
        '0203': 'ÚTILES ESCOLARES',
        '0204': 'DESODORANTES',
        '0205': 'JABONES DE BAÑO',
        '0206': 'SHAMPOO Y ACONDICIONADOR',
        '0207': 'MÁQUINAS DE AFEITAR',
        '0208': 'CREMAS Y CEPILLOS DENTALES',
        '0209': 'TOALLAS HIGIÉNICAS',
        '0210': 'ARTÍCULOS DE ASEO',
        '0211': 'DESECHABLES',
        'GA01': 'PASTILLERÍA Y SALUD',
        'GA02': 'ARTÍCULOS DE BEBÉ',
        'GA03': 'ÚTILES ESCOLARES',
        'GA04': 'DESODORANTES',
        'GA05': 'JABONES DE BAÑO',
        'GA06': 'SHAMPOO Y ACONDICIONADOR',
        'GA07': 'MÁQUINAS DE AFEITAR',
        'GA08': 'CREMAS Y CEPILLOS DENTALES',
        'GA09': 'TOALLAS HIGIÉNICAS',
        'GA10': 'ARTÍCULOS DE ASEO',
        'GA11': 'DESECHABLES',
        'GA12': 'LICORES Y BEBIDAS',
        'GA13': 'DULCERÍA Y COMESTIBLES',
        'GA14': 'ARTÍCULOS DE TOCADOR',
        'GA15': 'ARTÍCULOS PARA ZAPATERÍA',
        'GA16': 'ARTÍCULOS DE MODISTERÍA',
        'GA17': 'ARTÍCULOS DE PVC / PLÁSTICOS',
        'GA18': 'ARTÍCULOS ELÉCTRICOS',
        'GA19': 'VARIOS Y DIVERSOS',
        'GA20': 'ARTÍCULOS DE CABELLO',
        'GA21': 'VITRINAS',
        'GA22': 'MOLDES PARA LLAVES'
    }

    # Si existe reporte IN52 en la carpeta, extraer descripciones exactas del ERP
    item_in52_grupo = {}
    import glob
    in52_files = glob.glob(os.path.join(PROYECTO_DIR, 'IN52*.xlsx'))

    if in52_files:
        log.info(f"  -> Extrayendo mapeo de grupos desde {os.path.basename(in52_files[0])}...")
        try:
            import openpyxl, re
            wb_in52 = openpyxl.load_workbook(in52_files[0], data_only=True)
            ws_in52 = wb_in52.active
            for row in ws_in52.iter_rows(min_row=2, values_only=True):
                cod_itm = str(row[0] or '').strip()
                grp_raw = str(row[4] or '').strip()
                if grp_raw:
                    desc_raw = re.sub(r'^GA\d+\s*', '', grp_raw).strip().rstrip('-').strip()
                    desc_clean = clean_grupo_map.get(desc_raw, desc_raw)
                    item_in52_grupo[cod_itm] = desc_clean
            log.info(f"  -> Mapeados {len(item_in52_grupo)} ítems a sus Grupos GA desde IN52.")
        except Exception as e:
            log.error(f"Error al leer {in52_files[0]}: {e}")



    # 1. Cargar catálogo de productos (initem.dbf)
    log.info("Cargando catálogo de productos (initem.dbf)...")
    productos = {}
    total_descontinuados = 0

    for r in leer_dbf(DBF_ITEMS):
        cod = str(r.get('COD_ITEM', '')).strip()
        if not cod:
            cod = str(r.get('COD', '')).strip()
        if not cod:
            continue

        clase = str(r.get('CLASE', '')).strip().upper()
        if clase == 'X':
            total_descontinuados += 1
            continue

        descrip = str(r.get('DESCRIP', '')).strip()
        referencia = str(r.get('ITM_REF', '') or r.get('REFERENCIA', '') or cod).strip()

        # Determinar Grupo de Producto GA
        grupo_cod = str(r.get('GRUPO', '') or r.get('ITM_LINEA', '') or '').strip()
        if cod in item_in52_grupo:
            grupo_final = item_in52_grupo[cod]
        elif grupo_cod in clean_grupo_map:
            grupo_final = clean_grupo_map[grupo_cod]
        elif grupo_cod in grupos_ga_desc:
            grupo_final = grupos_ga_desc[grupo_cod]
        else:
            grupo_final = clean_grupo_map.get(grupo_cod, "VARIOS Y DIVERSOS")


        unimed = str(r.get('UNI_MEDIDA', '') or r.get('UNIMED', '') or 'UND').strip()
        minimo = safe_float(r.get('ITM_MINIMO', 0))
        maximo = safe_float(r.get('ITM_MAXIMO', 0))
        lista4 = safe_float(r.get('LISTA4', 0))
        lista5 = safe_float(r.get('LISTA5', 0))
        factor = safe_float(r.get('UNI_FACTOR', 1), default=1.0)

        productos[cod] = {
            'cod_item': cod,
            'referencia': referencia,
            'descrip': descrip,
            'code01': grupo_final,
            'unimed': unimed,
            'minimo': minimo,
            'maximo': maximo,
            'lista4': lista4,
            'lista5': lista5,
            'factor': factor
        }



    log.info(f"Catálogo cargado: {len(productos)} artículos activos ({total_descontinuados} excluidos por clase X).")

    # 2. Consolidar Saldos de Inventario
    log.info("Consolidando saldos de inventario desde ERP...")
    stock_actual = {}
    stock_loaded = False

    if os.path.exists(DBF_SALDOS):
        log.info(f"Leyendo saldos desde {os.path.basename(DBF_SALDOS)}...")
        try:
            for r in leer_dbf(DBF_SALDOS):
                cod = str(r.get('COD_SDO', '') or r.get('COD', '')).strip()
                cant = safe_float(r.get('ACTUAL_SDO', 0))
                if cod:
                    stock_actual[cod] = stock_actual.get(cod, 0.0) + cant
            stock_loaded = True
            log.info(f"Saldos cargados desde insaldo.dbf: {len(stock_actual)} registros.")
        except Exception as e:
            log.error(f"Error al leer insaldo.dbf: {e}")

    if not stock_loaded and os.path.exists(DBF_STOCK):
        log.info(f"Leyendo saldos desde {os.path.basename(DBF_STOCK)} (fallback)...")
        try:
            db_st = leer_dbf(DBF_STOCK)
            campos = [f.name.upper() for f in db_st.fields]
            campo_cod = next((c for c in ['COD', 'ST_COD', 'COD_ITEM'] if c in campos), campos[0])
            campo_cant = next((c for c in ['PIEZA', 'ST_PIEZA', 'CANT', 'STOCK', 'SALDO'] if c in campos), campos[-1])
            
            for r in db_st:
                cod = str(r.get(campo_cod, '')).strip()
                cant = safe_float(r.get(campo_cant, 0))
                if cod:
                    stock_actual[cod] = stock_actual.get(cod, 0.0) + cant
            log.info(f"Saldos cargados desde instock.dbf: {len(stock_actual)} registros.")
        except Exception as e:
            log.error(f"Error al leer instock.dbf: {e}")

    # 3. Cargar Movimientos de Ventas (inmvto.dbf) para Velocidad de Salida
    ventas_lineas = []
    if os.path.exists(DBF_MOVTOS):
        log.info("Procesando movimientos de ventas por producto (inmvto.dbf)...")
        for r in leer_dbf(DBF_MOVTOS):
            doc = str(r.get('MOV_DOC', '')).strip().upper()
            if doc not in ('FE', 'CT', 'N2', 'RM'):
                continue
            
            cod_prod = str(r.get('MOV_COD', '')).strip()
            prod_info = productos.get(cod_prod)
            if not prod_info:
                continue

            fecha_dt = r.get('MOV_FEC') or r.get('MOV_FECHA')
            if not fecha_dt:
                continue
            if isinstance(fecha_dt, datetime):
                fecha_dt = fecha_dt.date()
            if fecha_dt < FECHA_MIN or fecha_dt > FECHA_MAX:
                continue

            cant = safe_float(r.get('MOV_CANT', 0))
            if doc == 'N2':
                cant = -abs(cant)

            pve = safe_float(r.get('MOV_FC_PVE', 0))
            val_neto = safe_float(r.get('MOV_FC_VAL', 0))
            if val_neto <= 0:
                val_neto = cant * pve

            ventas_lineas.append({
                'ref': cod_prod,
                'articulo': prod_info['descrip'],
                'code01': prod_info['code01'],
                'fecha': obtener_serial_fecha(fecha_dt),
                'cant': cant,
                'valor': val_neto,
                'total1': val_neto
            })
        log.info(f"Cargados {len(ventas_lineas)} movimientos de venta válidos.")

    # 4. Generar Inventario.xlsx
    log.info("Generando Excel de Inventario (Inventario.xlsx)...")
    wb_inv = openpyxl.Workbook()
    ws_inv = wb_inv.active
    ws_inv.title = "Inventario"

    headers_inv = [
        "cod_item", "referencia", "articulo", "linea", "unimed",
        "stock_actual", "stock_min", "stock_max", "lista4", "lista5"
    ]
    ws_inv.append(headers_inv)

    for cod, prod in productos.items():
        st = stock_actual.get(cod, 0.0)
        ws_inv.append([
            prod['cod_item'],
            prod['referencia'],
            prod['descrip'],
            prod['code01'],
            prod['unimed'],
            st,
            prod['minimo'],
            prod['maximo'],
            prod['lista4'],
            prod['lista5']
        ])

    wb_inv.save(FILE_INVENTARIO)
    log.info(f"Guardado exitosamente: {FILE_INVENTARIO}")

    # 5. Generar Ventas por linea.xlsx
    log.info("Generando Excel de Ventas por Línea (Ventas por linea.xlsx)...")
    wb_vta = openpyxl.Workbook()
    ws_vta = wb_vta.active
    ws_vta.title = "Ventas por Linea"

    headers_vta = ["ref", "articulo", "code01", "fecha", "cant", "valor", "total1"]
    ws_vta.append(headers_vta)

    for row in ventas_lineas:
        ws_vta.append([
            row['ref'],
            row['articulo'],
            row['code01'],
            row['fecha'],
            row['cant'],
            row['valor'],
            row['total1']
        ])

    wb_vta.save(FILE_VENTAS)
    log.info(f"Guardado exitosamente: {FILE_VENTAS}")

    # 6. Actualizar last_update.json
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(FILE_LAST_UPDATE, 'w', encoding='utf-8') as f:
        json.dump({"last_update": now_str}, f, ensure_ascii=False, indent=2)

    # 7. Copiar automáticamente a la carpeta de producción de IIS si existe en el servidor
    IIS_DIR = r'C:\inetpub\wwwroot\tableroinventario'
    if os.path.exists(IIS_DIR):
        log.info(f"Copiando archivos actualizados a la carpeta de IIS ({IIS_DIR})...")
        try:
            shutil.copy2(FILE_INVENTARIO, os.path.join(IIS_DIR, 'Inventario.xlsx'))
            shutil.copy2(FILE_VENTAS, os.path.join(IIS_DIR, 'Ventas por linea.xlsx'))
            shutil.copy2(FILE_LAST_UPDATE, os.path.join(IIS_DIR, 'last_update.json'))
            log.info("Sincronización con IIS completada exitosamente.")
        except Exception as e:
            log.error(f"Error al copiar a la carpeta IIS: {e}")

    # También actualizar dist/ si existe en el entorno local
    DIST_DIR = os.path.join(PROYECTO_DIR, 'dist')
    if os.path.exists(DIST_DIR):
        try:
            shutil.copy2(FILE_INVENTARIO, os.path.join(DIST_DIR, 'Inventario.xlsx'))
            shutil.copy2(FILE_VENTAS, os.path.join(DIST_DIR, 'Ventas por linea.xlsx'))
            shutil.copy2(FILE_LAST_UPDATE, os.path.join(DIST_DIR, 'last_update.json'))
        except Exception:
            pass

    log.info(f"Proceso finalizado con éxito a las {now_str}.")

if __name__ == '__main__':
    main()

