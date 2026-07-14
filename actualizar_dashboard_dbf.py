# =============================================================================
# SCRIPT AUTOMATICO - Actualizacion Directa del Dashboard desde ERP (DBF)
# Uso: python actualizar_dashboard_dbf.py
# Programar en el Programador de Tareas de Windows para correr diariamente
# =============================================================================

import os
import sys
import logging
import shutil
import json
from datetime import datetime, date

# Intentar importar librerias requeridas o instalarlas automaticamente
try:
    from dbfread import DBF
except ImportError:
    print("Instalando libreria dbfread...")
    os.system('pip install dbfread')
    from dbfread import DBF

try:
    import openpyxl
except ImportError:
    print("Instalando libreria openpyxl...")
    os.system('pip install openpyxl')
    import openpyxl

# =============================================================================
# CONFIGURACION
# =============================================================================
RUTA_DATOS     = r'E:\DataX_NUEVO\datos'   # Ruta a los archivos DBF del ERP
# Obtener la ruta del proyecto dinamicamente donde se encuentre el script
PROYECTO_DIR   = os.path.dirname(os.path.abspath(__file__))

# Fallback para pruebas locales si no existe la ruta en el servidor
if not os.path.exists(RUTA_DATOS):
    fallbacks = [
        r"D:\4 Hector Fabio\Distribuidora JR\Base de datos",
        os.path.join(os.path.dirname(PROYECTO_DIR), 'Distribuidora JR', 'Base de datos')
    ]
    for fb in fallbacks:
        if os.path.exists(fb):
            RUTA_DATOS = fb
            break

# Archivos DBF de origen
DBF_ITEMS      = os.path.join(RUTA_DATOS, 'initem.dbf')   # Catalogo de productos
DBF_FACT       = os.path.join(RUTA_DATOS, 'infact.dbf')   # Cabecera de facturas
DBF_MOVTOS     = os.path.join(RUTA_DATOS, 'inmvto.dbf')   # Detalles de movimientos/ventas
DBF_VENDEDORES = os.path.join(RUTA_DATOS, 'cgvend.dbf')   # Maestro de vendedores
DBF_CLIENTES   = os.path.join(RUTA_DATOS, 'cgbenf.dbf')   # Maestro de clientes
DBF_CIUDADES   = os.path.join(RUTA_DATOS, 'cgciudad.dbf') # Maestro de ciudades
DBF_SALDOS_CARTERA = os.path.join(RUTA_DATOS, 'cgsaldo.dbf') # Saldos de cartera

# Nombres de archivos del Dashboard
FILE_MAESTRA   = '1Maestra de clientes2026.xlsx'
FILE_VENTAS    = 'Ventas por linea.xlsx'
FILE_INVENTARIO = 'Inventario.xlsx'
FILE_CARTERA   = 'Cartera.xlsx'

# Rango de fechas: Desde Septiembre 2025 hasta el dia de hoy (dinamico)
FECHA_MIN      = date(2025, 9, 1)
FECHA_MAX      = date.today()

# Filtros de Documentos (basado en su configuracion)
DOCS_VENTA      = ('FE', 'CT')
DOCS_DEVOLUCION = ('N2',)

# =============================================================================
# INICIALIZACION DE LOGS
# =============================================================================
LOG_FILE = os.path.join(PROYECTO_DIR, 'log_dbf_dashboard.txt')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger()

from dbfread.field_parser import FieldParser

class SafeFieldParser(FieldParser):
    def parseN(self, field, data):
        try:
            data_clean = data.replace(b'\x00', b'').strip()
            if not data_clean:
                return 0
            return super().parseN(field, data_clean)
        except ValueError:
            return 0

    def parseF(self, field, data):
        try:
            data_clean = data.replace(b'\x00', b'').strip()
            if not data_clean:
                return 0.0
            return super().parseF(field, data_clean)
        except ValueError:
            return 0.0

def leer_dbf(ruta):
    return DBF(ruta, load=False, encoding='latin-1', ignore_missing_memofile=True, parserclass=SafeFieldParser)

def obtener_nombre_mes(mes_num):
    meses = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    if 1 <= mes_num <= 12:
        return meses[mes_num - 1]
    return 'Desconocido'

def obtener_serial_fecha(d):
    if d is None:
        return 0
    if isinstance(d, datetime):
        d = d.date()
    elif not isinstance(d, date):
        return 0
    fecha_base = date(1899, 12, 30)
    return (d - fecha_base).days

def safe_float(val, default=0.0):
    try:
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).strip()
        if not val_str:
            return default
        val_str = val_str.replace(',', '.')
        return float(val_str)
    except (ValueError, TypeError):
        return default

# Helper para buscar campos de forma insensible a mayusculas/minusculas
def get_field_val(row, keys_to_try, default=''):
    row_keys = {k.upper(): k for k in row.keys()}
    for k in keys_to_try:
        ku = k.upper()
        if ku in row_keys:
            val = row[row_keys[ku]]
            if val is not None:
                return val
    return default

def exportar_datos():
    log.info('=== INICIO PROCESO DE ACTUALIZACION DEL DASHBOARD ===')
    log.info(f'Procesando rango de fechas: {FECHA_MIN} al {FECHA_MAX}')

    # 1. Cargar Maestro de Vendedores (cgvend.dbf)
    log.info('Cargando maestro de vendedores...')
    vendedores = {}
    if os.path.exists(DBF_VENDEDORES):
        for r in leer_dbf(DBF_VENDEDORES):
            cod = str(r.get('COD_VEND', '')).strip()
            des = str(r.get('DES_VEND', '')).strip()
            if cod:
                vendedores[cod] = des
        log.info(f'  -> Cargados {len(vendedores)} vendedores.')
    else:
        log.warning(f'  -> No se encontro {DBF_VENDEDORES}. Se usaran codigos por defecto.')

    # 2. Cargar Maestro de Clientes (cgbenf.dbf)
    log.info('Cargando maestro de clientes...')
    clientes_master = {}
    if os.path.exists(DBF_CLIENTES):
        for r in leer_dbf(DBF_CLIENTES):
            cod = str(r.get('COD_BENF', '')).strip()
            nom = str(r.get('NOM_BENF', '')).strip()
            nit = str(r.get('NIT_BENF', '')).strip()
            cupo = safe_float(r.get('CUP_BENF', 0.0))
            cod_vend = str(r.get('VENDEDOR_B', '')).strip()
            if cod:
                clientes_master[cod] = {
                    'nombre': nom,
                    'nit': nit,
                    'cupo': cupo,
                    'vendedor_maestro': cod_vend
                }
        log.info(f'  -> Cargados {len(clientes_master)} clientes.')
    else:
        log.warning(f'  -> No se encontro {DBF_CLIENTES}. Se usaran datos de facturacion directos.')

    # 3. Intentar buscar Maestro de Lineas (inlinea.dbf / cglinea.dbf)
    log.info('Buscando catalogo de lineas comerciales...')
    lineas_desc = {
        '0000': 'Linea General',
        '0001': 'SUPER DE ALIMENTOS S.A',
        '0002': 'CONFITECA COLOMBIA S.A',
        '0003': 'JOHNSON & JOHNSON',
        '0004': 'TECNOQUIMICAS S.A',
        '0005': 'ETERNA S.A',
        '0006': 'MARCHEN S.A',
        '0009': 'BEISBOL DE COLOMBIA S.A',
        '0010': 'JGB S.A',
        '0011': 'LABORATORIOS OSA S.A',
        '0012': 'BAYER S.A',
        '0013': 'DRYPERS ANDINA S.A',
        '0014': 'PFIZER S.A',
        '0015': 'MULTIDIMENSIONALES S.A',
        '0016': 'BELLEZA EXPRESS S.A',
        '0017': 'WA S.A',
        '0018': 'SC JOHNSON',
        '0019': 'BIC COLOMBIA S.A',
        '0020': 'HADA S.A',
        '0021': 'FAMILY HOME PRODUCTS S.A',
        '0022': 'QUALA S.A',
        '0023': 'PROCTER & GAMBLE',
        '0024': 'COMESTIBLES ITALO S.A',
        '0025': 'UNIBOL S.A.S',
        '0026': 'GRADEZCO LTDA',
        '0027': 'VELAS Y VELONES SAN JORGE',
        '0028': 'COLOMBIANA KIMBERLY COLPAPEL S.A',
        '0029': 'ENERGIZER DE COLOMBIA S.A',
        '0031': 'NACIONAL DE PILAS OCCIDENTE S.A',
        '0032': 'LABORATORIO BACHUE S.A.S',
        '0033': 'UNILEVER ANDINA',
        '0034': 'OFFI-ESCO',
        '0035': 'EDGEWELL PERSONAL CARE',
        '0036': 'COMERCIALIZADORA COMARKO SAS',
        '0037': 'AGRICOLA HIMALAYA S.A.',
        '0038': 'HALEON',
        '0039': 'LABORATORIO LIMAR SAS',
        '0040': 'HARTUNG & COMPANIA SAS',
        '0041': 'KAALPANIK COLOMBIA SAS VATIO',
        '0042': 'COLGATE PALMOLIVE'
    }
    line_dbfs_to_try = ['inlinea.dbf', 'cglinea.dbf', 'inlin.dbf']
    dbf_lineas_path = None
    for f in line_dbfs_to_try:
        p = os.path.join(RUTA_DATOS, f)
        if os.path.exists(p):
            dbf_lineas_path = p
            break
            
    if dbf_lineas_path:
        log.info(f'Cargando catalogo de lineas desde {os.path.basename(dbf_lineas_path)}...')
        for r in leer_dbf(dbf_lineas_path):
            cod = str(get_field_val(r, ['COD_LINEA', 'CODE_LIN', 'COD', 'B2_CODE'])).strip()
            des = str(get_field_val(r, ['NOM_LINEA', 'DES_LINEA', 'DESCRIP', 'B2_DESC'])).strip()
            if cod:
                lineas_desc[cod] = des
        log.info(f'  -> Cargadas {len(lineas_desc)} descripciones de linea.')
    else:
        log.warning('  -> No se encontro catalogo de lineas. Se usaran codigos de linea por defecto.')

    # 3.5 Cargar Precios de Lista 4 y Lista 5 para costos de ultima compra
    lista4_precios = {}
    lista5_precios = {}
    dbf_lista_path = None
    if os.path.exists(RUTA_DATOS):
        for f in os.listdir(RUTA_DATOS):
            if f.lower() == 'inlista.dbf':
                dbf_lista_path = os.path.join(RUTA_DATOS, f)
                break
                
    if dbf_lista_path:
        log.info(f'Cargando precios de Lista 4 y Lista 5 desde {os.path.basename(dbf_lista_path)}...')
        try:
            db = leer_dbf(dbf_lista_path)
            campos = [f.name.upper() for f in db.fields]
            campo_cod = None
            for c in ['COD_LIS', 'COD_ITEM', 'COD', 'ITEM']:
                if c in campos:
                    campo_cod = c
                    break
            if not campo_cod:
                campo_cod = campos[0]
                
            campo_precio4 = 'PRECIO4' if 'PRECIO4' in campos else None
            if not campo_precio4:
                for c in campos:
                    if 'PRECIO4' in c or 'LIS4' in c:
                        campo_precio4 = c
                        break
            if not campo_precio4:
                campo_precio4 = 'PRECIO4'
                
            campo_precio5 = 'PRECIO5' if 'PRECIO5' in campos else None
            if not campo_precio5:
                for c in campos:
                    if 'PRECIO5' in c or 'LIS5' in c:
                        campo_precio5 = c
                        break
            if not campo_precio5:
                campo_precio5 = 'PRECIO5'
                
            log.info(f"  -> Usando campos de Lista: codigo='{campo_cod}', lista4='{campo_precio4}', lista5='{campo_precio5}'")
            
            # Cargar en memoria para poder recorrerlo multiples veces sin agotar el generador
            lista_records = list(db)
                
            for r in lista_records:
                cod = str(r.get(campo_cod, '')).strip()
                if cod:
                    try:
                        p4_val = safe_float(r.get(campo_precio4, 0))
                        lista4_precios[cod] = p4_val
                    except (ValueError, TypeError):
                        pass
                    try:
                        p5_val = safe_float(r.get(campo_precio5, 0))
                        lista5_precios[cod] = p5_val
                    except (ValueError, TypeError):
                        pass
            log.info(f"  -> Cargados {len(lista4_precios)} precios de Lista 4 y {len(lista5_precios)} de Lista 5.")
        except Exception as e:
            log.error(f'Error al leer inlista.dbf: {e}')
    else:
        log.warning('  -> No se encontro inlista.dbf. Se usaran fallbacks por defecto.')

    # 4. Cargar Catalogo de Productos (initem.dbf)
    log.info('Cargando catalogo de productos...')
    productos = {}
    for r in leer_dbf(DBF_ITEMS):
        cod = str(r.get('COD_ITEM', '')).strip()
        if cod:
            linea_cod = str(r.get('ITM_LINEA', '')).strip()
            
            # Formatear el campo code01 del Excel: "CODIGO NOMBRE_LINEA"
            linea_nom = lineas_desc.get(linea_cod, 'Sin Descripcion de Linea')
            code01 = f"{linea_cod} {linea_nom}".strip()
            if not linea_cod:
                code01 = "0000 Sin Descripcion de Linea"
                
            # Determinar costo: usar precio de Lista 4 si existe, sino fallback a ULT_COSTO
            costo_lista4 = lista4_precios.get(cod)
            if costo_lista4 is not None:
                costo_final = costo_lista4
            else:
                costo_final = safe_float(r.get('ULT_COSTO', 0))

            costo_lista5 = lista5_precios.get(cod, 0.0)

            productos[cod] = {
                'descrip'  : str(r.get('DESCRIP',  '')).strip(),
                'code01'   : code01,
                'linea'    : linea_cod,
                'minimo'   : safe_float(r.get('ITM_MINIMO', 0)),
                'maximo'   : safe_float(r.get('ITM_MAXIMO', 0)),
                'unimed'   : str(r.get('UNIMED',   '')).strip(),
                'clase'    : str(r.get('CLASE',    '')).strip().upper(),
                'lista4'   : costo_final,
                'lista5'   : costo_lista5,
                'referencia': str(r.get('REFERENCIA', '')).strip(),
                # Intentamos leer factor de conversion de empaque si existe
                'factor'   : float(get_field_val(r, ['UNI_FACTOR', 'FACTOR', 'CONV', 'EMP_CANT'], 1) or 1)
            }
    log.info(f'  -> Cargados {len(productos)} productos.')

    # 5. Cargar Cabeceras de Facturas (infact.dbf)
    log.info('Cargando facturas y armando Maestra de Clientes...')
    facturas = {}
    clientes_maestra = []
    claves_maestra_unicas = set()

    if not os.path.exists(DBF_FACT):
        log.warning(f"No se encontro {DBF_FACT}. Se saltara la carga de facturas.")
        dbf_fact_records = []
    else:
        dbf_fact_records = leer_dbf(DBF_FACT)

    for r in dbf_fact_records:
        doc = str(r.get('FC_DOC', '')).strip()
        if not doc:
            continue
            
        # Validar combinaciones validas de Centro de Utilidad (Sucursal) y Tipo de Documento
        cu = str(r.get('FC_CU', '')).strip()
        if cu == '01' and doc not in ('FE', 'N2'):
            continue
        elif cu == '02' and doc not in ('CT',):
            continue
        elif cu not in ('01', '02'):
            continue
            
        # Omitir anuladas
        if r.get('FC_ANULA'):
            continue
            
        fecha = r.get('FC_FECHA')
        if not fecha or fecha < FECHA_MIN or fecha > FECHA_MAX:
            continue

        nro = str(r.get('FC_NRO', '')).strip()
        docto = f"{doc}-{nro}" # Formato: FE-XXXXXX o CT-XXXXXX
        
        cod_vend = str(r.get('FC_VENTAS', '')).strip()
        nom_vend = vendedores.get(cod_vend, 'PRINCIPAL' if cod_vend == '01' else 'VENDEDOR')
        vendedor_completo = f"{cod_vend} {nom_vend}".strip()
        
        cod_client = str(r.get('FC_BENF', '')).strip()
        
        # Enriquecer nombre de cliente desde el maestro de clientes
        info_c = clientes_master.get(cod_client, {})
        nombre_cli = info_c.get('nombre', str(r.get('FC_RAZON', '')).strip())
        
        # Calcular total a partir de los campos reales de infact.dbf
        vlr_bruto = safe_float(r.get('FC_VLRBRUT', 0))
        vlr_desc = safe_float(r.get('FC_VLR_DSC', 0))
        vlr_iva = safe_float(r.get('FC_VLR_IVA', 0))
        total1 = vlr_bruto - vlr_desc + vlr_iva
        facturas[docto] = {
            'vendedor'  : vendedor_completo,
            'cod_client': cod_client,
            'nombre_cli': nombre_cli,
            'fecha'     : fecha,
            'mes'       : obtener_nombre_mes(fecha.month),
            'tipo'      : doc,
            'nro'       : nro,
        }

        # Guardar registro para 1Maestra de clientes2026.xlsx
        key_maestra = (docto, vendedor_completo, cod_client)
        if key_maestra not in claves_maestra_unicas:
            claves_maestra_unicas.add(key_maestra)
            
            # Formato benf1: "NOMBRE_CLIENTE COD_CLIENTE"
            benf1 = f"{nombre_cli} {cod_client}".strip()
            
            clientes_maestra.append({
                'docto': docto,
                'fecha': obtener_serial_fecha(fecha),
                'Mes': obtener_nombre_mes(fecha.month),
                'benf1': benf1,
                'total1': total1,
                'vendedor': vendedor_completo,
                'cod_client': cod_client,
                'nombre_cli': nombre_cli,
                'Tipo': doc
            })

    # =========================================================================
    # 6. Cargar Detalles de Ventas (inmvto.dbf) con filtro Consumidor Final
    # =========================================================================
    def calcular_valor_movimiento(r):
        cant = safe_float(r.get('MOV_CANT', 0))
        pve = safe_float(r.get('MOV_FC_PVE', 0))
        dsc1 = safe_float(r.get('MOV_FC_DSC', 0))
        dsc2 = safe_float(r.get('MOV_FC_DS2', 0))
        iva_pct = safe_float(r.get('MOV_FC_IMP', 0))
        
        val_neto_dbf = safe_float(r.get('MOV_FC_VAL', 0))
        
        if val_neto_dbf > 100:
            val = val_neto_dbf
            if iva_pct <= 100:
                iva_val = val * (iva_pct / 100.0)
            else:
                iva_val = iva_pct
            total = val + iva_val
        else:
            subtotal = cant * pve
            if 0 < dsc1 <= 100:
                subtotal *= (1.0 - dsc1 / 100.0)
            elif dsc1 > 100:
                subtotal = max(0.0, subtotal - dsc1)
                
            if 0 < dsc2 <= 100:
                subtotal *= (1.0 - dsc2 / 100.0)
            elif dsc2 > 100:
                subtotal = max(0.0, subtotal - dsc2)
                
            val = subtotal
            iva_val = val * (iva_pct / 100.0)
            total = val + iva_val
            
        return val, iva_val, total

    log.info('Procesando movimientos de ventas por linea...')
    ventas_lineas = []
    
    # Pre-cargar Ajustes de Entrada (Concepto 30, Doc AJ) para filtro Consumidor Final
    ajustes_entrada = {}
    if os.path.exists(DBF_MOVTOS):
        log.info('Pre-escaneando ajustes de entrada en inmvto.dbf...')
        for r in leer_dbf(DBF_MOVTOS):
            doc = str(r.get('MOV_DOC', '')).strip().upper()
            cpto = str(r.get('MOV_CPTO', '')).strip()
            if doc == 'AJ' and cpto == '30':
                fecha = r.get('MOV_FEC')
                cod = str(r.get('MOV_COD', '')).strip()
                cant = safe_float(r.get('MOV_CANT', 0))
                if fecha and cod and cant > 0:
                    key = (fecha, cod, cant)
                    ajustes_entrada[key] = ajustes_entrada.get(key, 0) + 1
        log.info(f'  -> Encontrados {sum(ajustes_entrada.values())} ajustes de entrada.')

        log.info('Procesando movimientos de ventas...')
        total_compensado = {} # key: docto, value: float (amount subtracted)
        excluidos_lineas_count = 0
        
        for r in leer_dbf(DBF_MOVTOS):
            doc = str(r.get('MOV_DOC', '')).strip().upper()
            if doc not in ('FE', 'CT', 'N2', 'RM'):
                continue
            
            nro = str(r.get('MOV_NRO', '')).strip()
            
            # Determinar cabecera (factura) asociada
            fact_hdr = None
            if doc == 'RM':
                facrem = str(r.get('MOV_FACREM', '')).strip().upper()
                if not facrem:
                    continue
                
                # Intentar extraer tipo y numero de factura de MOV_FACREM
                target_doc = None
                target_nro = None
                for prefix in ['FE', 'CT', 'N2']:
                    if facrem.startswith(prefix):
                        target_doc = prefix
                        target_nro = facrem[len(prefix):].strip()
                        break
                
                if target_doc and target_nro:
                    fact_hdr = facturas.get(f"{target_doc}-{target_nro}")
                    if not fact_hdr:
                        fact_hdr = facturas.get(f"{target_doc}-{target_nro.lstrip('0') or '0'}")
                else:
                    target_doc = 'FE'
                    target_nro = facrem
                    fact_hdr = facturas.get(f"FE-{facrem}")
                    if not fact_hdr:
                        fact_hdr = facturas.get(f"FE-{facrem.lstrip('0') or '0'}")
            else:
                fact_hdr = facturas.get(f"{doc}-{nro}")
                if not fact_hdr:
                    fact_hdr = facturas.get(f"{doc}-{nro.lstrip('0') or '0'}")
            
            if not fact_hdr:
                continue
                
            docto = f"{fact_hdr['tipo']}-{fact_hdr['nro']}"
            fecha_dt = fact_hdr['fecha']
            cod_client = fact_hdr['cod_client']
            
            cod_prod = str(r.get('MOV_COD', '')).strip()
            prod_info = productos.get(cod_prod)
            if not prod_info:
                continue
                
            if not fecha_dt or fecha_dt < FECHA_MIN or fecha_dt > FECHA_MAX:
                continue
                
            cant = safe_float(r.get('MOV_CANT', 0))

            # FILTRO: Ajuste Netted (Consumidor Final)
            # Concepto 30, same-day entry adjustment
            is_fictitious = False
            if cod_client == '222222222222' and cant > 0:
                adj_key = (fecha_dt, cod_prod, cant)
                if ajustes_entrada.get(adj_key, 0) > 0:
                    ajustes_entrada[adj_key] -= 1
                    is_fictitious = True
                    excluidos_lineas_count += 1
                    
                    # Calcular valor compensado
                    val_neto, val_iva, total_neto = calcular_valor_movimiento(r)
                    
                    total_compensado[docto] = total_compensado.get(docto, 0.0) + total_neto
                    continue # Excluir la linea

            # Si no es ficticio, se procesa la venta
            vendedor = fact_hdr['vendedor']
            code01 = prod_info['code01']
            articulo = prod_info['descrip']
            cliente = fact_hdr['nombre_cli']
            
            val, iva, total1 = calcular_valor_movimiento(r)
            
            detalle = str(r.get('MOV_DETALL', '')).strip()
            observa = ''
            remision = str(r.get('MOV_FACREM', '')).strip()
            if not remision:
                remision = '   -'
                
            q_adi = cant
            n_adi = prod_info['unimed']
            uni_factor = prod_info['factor']
            emp = '0202'
            
            ventas_lineas.append({
                'vendedor': vendedor,
                'code01': code01,
                'docto': docto,
                'fecha': obtener_serial_fecha(fecha_dt),
                'ref': cod_prod,
                'articulo': articulo,
                'cant': cant,
                'cliente': cliente,
                'valor': val,
                'iva': iva,
                'total1': total1,
                'detalle': detalle,
                'observa': observa,
                'remision': remision,
                'q_adi': q_adi,
                'n_adi': n_adi,
                'uni_factor': uni_factor,
                'cod_benf': cod_client,
                'emp': emp
            })
            
        # Ajustar clientes_maestra con los valores compensados
        clientes_maestra_ajustada = []
        facturas_modificadas_count = 0
        facturas_eliminadas_count = 0
        
        for cli in clientes_maestra:
            doc = cli['docto']
            if doc in total_compensado:
                compensado = total_compensado[doc]
                nuevo_total = cli['total1'] - compensado
                if nuevo_total <= 1.0: # Si cae a cero o cercano a cero, se elimina
                    facturas_eliminadas_count += 1
                    continue
                else:
                    cli['total1'] = nuevo_total
                    facturas_modificadas_count += 1
            clientes_maestra_ajustada.append(cli)
            
        clientes_maestra = clientes_maestra_ajustada
        log.info(f"  -> Procesados {len(clientes_maestra)} registros para Maestra de Clientes.")
        log.info(f"Filtro Consumidor Final: Excluidas {excluidos_lineas_count} lineas de venta.")
        log.info(f"Filtro Consumidor Final: Ajustadas {facturas_modificadas_count} facturas y eliminadas {facturas_eliminadas_count} facturas.")
    else:
        log.warning(f"No se encontro {DBF_MOVTOS}. Se saltara la carga de movimientos.")

    # =========================================================================
    # 7. Consolidar saldos de inventario desde el ERP
    # =========================================================================
    log.info('Consolidando saldos de inventario desde el ERP...')
    stock_actual = {}
    
    DBF_SALDOS = os.path.join(RUTA_DATOS, 'insaldo.dbf')
    DBF_STOCK_FILE = os.path.join(RUTA_DATOS, 'instock.dbf')
    
    loaded_stock = False
    if os.path.exists(DBF_SALDOS):
        log.info(f"  -> Leyendo stock desde {os.path.basename(DBF_SALDOS)}...")
        try:
            for r in leer_dbf(DBF_SALDOS):
                cod = str(r.get('COD_SDO', '')).strip()
                cant = safe_float(r.get('ACTUAL_SDO', 0))
                if cod:
                    stock_actual[cod] = stock_actual.get(cod, 0.0) + cant
            log.info(f"  -> Cargados {len(stock_actual)} saldos desde insaldo.dbf.")
            loaded_stock = True
        except Exception as e:
            log.error(f"Error al leer insaldo.dbf: {e}")
            
    if not loaded_stock and os.path.exists(DBF_STOCK_FILE):
        log.info(f"  -> Leyendo stock desde {os.path.basename(DBF_STOCK_FILE)} (fallback)...")
        try:
            db_st = leer_dbf(DBF_STOCK_FILE)
            campos = [f.name.upper() for f in db_st.fields]
            campo_cod = next((c for c in ['COD', 'ST_COD', 'COD_ITEM'] if c in campos), campos[0])
            campo_cant = next((c for c in ['PIEZA', 'ST_PIEZA', 'CANT', 'STOCK', 'SALDO', 'EXI'] if c in campos), None)
            if not campo_cant:
                for c in campos:
                    if any(x in c for x in ['CANT', 'STOCK', 'PIEZA', 'SALDO', 'EXI']):
                        campo_cant = c
                        break
            if not campo_cant:
                campo_cant = campos[-1]
                
            log.info(f"  -> Usando campos de Stock: codigo='{campo_cod}', cantidad='{campo_cant}'")
            for r in db_st:
                cod = str(r.get(campo_cod, '')).strip()
                cant = safe_float(r.get(campo_cant, 0))
                if cod:
                    stock_actual[cod] = stock_actual.get(cod, 0.0) + cant
            log.info(f"  -> Cargados {len(stock_actual)} saldos desde instock.dbf.")
            loaded_stock = True
        except Exception as e:
            log.error(f"Error al leer instock.dbf: {e}")
            
    if not loaded_stock or len(stock_actual) == 0:
        log.info("  -> No se obtuvieron saldos de stock desde DBF. Buscando reporte de saldos manual SaldosInv_*.xlsx...")
        import glob
        saldos_files = glob.glob(os.path.join(PROYECTO_DIR, 'SaldosInv_*.xlsx'))
        if saldos_files:
            saldos_files.sort(reverse=True)
            manual_file = saldos_files[0]
            log.info(f"  -> Leyendo saldos de stock alternativos desde {os.path.basename(manual_file)}...")
            try:
                wb_m = openpyxl.load_workbook(manual_file, data_only=True)
                ws_m = wb_m.active
                rows_m = list(ws_m.iter_rows(values_only=True))
                h = rows_m[0]
                idx_cod = next((i for i, x in enumerate(h) if x and 'cod_item' in str(x).lower()), 0)
                idx_saldo = next((i for i, x in enumerate(h) if x and '_saldo' in str(x).lower()), 9)
                
                for r in rows_m[1:]:
                    if len(r) > max(idx_cod, idx_saldo):
                        cod = str(r[idx_cod]).strip()
                        cant = float(r[idx_saldo] or 0)
                        if cod:
                            stock_actual[cod] = stock_actual.get(cod, 0.0) + cant
                log.info(f"  -> Cargados {len(stock_actual)} saldos desde el Excel manual.")
                loaded_stock = True
            except Exception as e:
                log.error(f"Error al leer Excel manual de saldos: {e}")
        else:
            log.warning("  -> No se encontraron archivos de saldos manuales SaldosInv_*.xlsx.")

    inventario_rows = []
    for cod, prod in productos.items():
        # EXCLUIR PRODUCTOS CON CLASE == 'X' (descontinuados/inactivos)
        if prod['clase'] == 'X':
            continue
        stock = stock_actual.get(cod, 0.0)
        inventario_rows.append({
            'cod_item': cod,
            'referencia': prod['referencia'] or cod,
            'articulo': prod['descrip'],
            'linea': prod['code01'],
            'unimed': prod['unimed'] or 'UND',
            'stock_actual': stock,
            'stock_min': prod['minimo'],
            'stock_max': prod['maximo'],
            'lista4': prod['lista4'],
            'lista5': prod['lista5']
        })

    # =========================================================================
    # 7.5 Procesar Cartera (cgsaldo.dbf)
    # =========================================================================
    log.info('Procesando cartera (cgsaldo.dbf)...')
    cartera_clientes = {}
    cartera_documentos = []
    today = date.today()

    if os.path.exists(DBF_SALDOS_CARTERA):
        registros_procesados = 0
        saldos_activos = 0
        for r in leer_dbf(DBF_SALDOS_CARTERA):
            registros_procesados += 1
            cuenta = str(r.get('CUENTA_SAL', '')).strip()
            # Filtrar por cuenta de cartera comercial (habitualmente 1305*)
            if not cuenta.startswith('1305'):
                continue
                
            deb = safe_float(r.get('DEB_SALDO', 0.0))
            cre = safe_float(r.get('CRE_SALDO', 0.0))
            saldo = deb - cre
            
            # Solo facturas con saldo pendiente
            if saldo <= 0.01:
                continue
                
            saldos_activos += 1
            benf = str(r.get('BENF_SALDO', '')).strip()
            doc_tipo = str(r.get('DOC_SALDO', '')).strip()
            doc_nro = str(r.get('NRO_SALDO', '')).strip()
            documento = f"{doc_tipo}-{doc_nro}"
            
            fec_elab = r.get('FECELA_SAL')
            fec_vcto = r.get('FECVCT_SAL')
            cod_vend = str(r.get('VEND_SALDO', '')).strip()
            nombre_vend = vendedores.get(cod_vend, 'Sin Vendedor')
            
            # Calcular dias de mora
            dias_mora = 0
            estado = 'Corriente'
            if fec_vcto:
                if isinstance(fec_vcto, datetime):
                    fec_vcto = fec_vcto.date()
                elif isinstance(fec_vcto, date):
                    pass
                else:
                    fec_vcto = None
                
                if fec_vcto and fec_vcto < today:
                    dias_mora = (today - fec_vcto).days
                    estado = 'Vencido'
                    
            # Obtener datos del cliente del maestro
            info_cli = clientes_master.get(benf, {'nombre': 'Cliente Desconocido', 'nit': benf, 'cupo': 0.0, 'vendedor_maestro': ''})
            nombre_cli = info_cli['nombre']
            nit_cli = info_cli['nit']
            
            # 1. Registrar documento
            cartera_documentos.append({
                'cod_benf': benf,
                'nit': nit_cli,
                'nombre': nombre_cli,
                'documento': documento,
                'fecha_elab': obtener_serial_fecha(fec_elab),
                'fecha_vcto': obtener_serial_fecha(fec_vcto),
                'dias_mora': dias_mora,
                'valor_original': deb,
                'abono': cre,
                'saldo': saldo,
                'estado': estado,
                'cod_vend': cod_vend,
                'nombre_vend': nombre_vend
            })
            
            # 2. Acumular saldo del cliente
            if benf not in cartera_clientes:
                cod_vend_cli = info_cli.get('vendedor_maestro', cod_vend) or cod_vend
                nombre_vend_cli = vendedores.get(cod_vend_cli, 'Sin Vendedor')
                cartera_clientes[benf] = {
                    'cod_benf': benf,
                    'nit': nit_cli,
                    'nombre': nombre_cli,
                    'cupo_asignado': info_cli['cupo'],
                    'saldo_total': 0.0,
                    'saldo_corriente': 0.0,
                    'saldo_vencido': 0.0,
                    'cupo_disponible': info_cli['cupo'],
                    'mora_maxima': 0,
                    'cod_vend': cod_vend_cli,
                    'nombre_vend': nombre_vend_cli
                }
                
            cartera_clientes[benf]['saldo_total'] += saldo
            if estado == 'Vencido':
                cartera_clientes[benf]['saldo_vencido'] += saldo
                cartera_clientes[benf]['mora_maxima'] = max(cartera_clientes[benf]['mora_maxima'], dias_mora)
            else:
                cartera_clientes[benf]['saldo_corriente'] += saldo
                
            cartera_clientes[benf]['cupo_disponible'] = cartera_clientes[benf]['cupo_asignado'] - cartera_clientes[benf]['saldo_total']

        log.info(f"  -> Procesados {registros_procesados} registros de cgsaldo.dbf.")
        log.info(f"  -> Encontrados {saldos_activos} documentos con saldo activo.")
        log.info(f"  -> Consolidados {len(cartera_clientes)} clientes en cartera.")
    else:
        log.warning(f"  -> No se encontro la tabla de cartera {DBF_SALDOS_CARTERA}.")

    # =========================================================================
    # 8. Guardar archivos Excel
    # =========================================================================
    def guardar_excel(ruta, sheet_name, headers, data):
        log.info(f"Guardando archivo: {os.path.basename(ruta)}...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Escribir encabezados
        ws.append(headers)
        
        # Escribir filas
        for row in data:
            row_data = [row.get(h, '') for h in headers]
            ws.append(row_data)
            
        wb.save(ruta)
        log.info(f"  -> Guardado en: {ruta}")

    def guardar_excel_multisheet(ruta, sheets):
        log.info(f"Guardando archivo multidiseño: {os.path.basename(ruta)}...")
        wb = openpyxl.Workbook()
        # Eliminar la primera hoja por defecto
        if wb.active:
            wb.remove(wb.active)
            
        for sheet_name, headers, data in sheets:
            ws = wb.create_sheet(title=sheet_name)
            # Escribir encabezados
            ws.append(headers)
            # Escribir filas
            for row in data:
                row_data = [row.get(h, '') for h in headers]
                ws.append(row_data)
                
        wb.save(ruta)
        log.info(f"  -> Guardado en: {ruta}")

    path_maestra = os.path.join(PROYECTO_DIR, FILE_MAESTRA)
    path_ventas = os.path.join(PROYECTO_DIR, FILE_VENTAS)
    path_inventario = os.path.join(PROYECTO_DIR, FILE_INVENTARIO)
    path_cartera = os.path.join(PROYECTO_DIR, FILE_CARTERA)
    
    # Guardar en raiz
    if clientes_maestra and vendedores:
        guardar_excel(path_maestra, 'Maestra de Clientes', 
                      ['docto', 'fecha', 'Mes', 'benf1', 'total1', 'vendedor', 'cod_client', 'nombre_cli', 'Tipo'], 
                      clientes_maestra)
    else:
        log.warning(f"  -> No hay registros para Maestra de Clientes o no se cargo cgvend.dbf. Se omite la sobrescritura de {FILE_MAESTRA}.")
        
    # DIAGNOSTICO DE VENTAS PARA ARTICULO 06766 (2200)
    diag_ventas_2200 = [v for v in ventas_lineas if str(v.get('ref', '')).strip().lstrip('0') in ('6766', '2200')]
    log.info(f"  -> [DIAGNOSTICO 2200] Registros en ventas_lineas: {len(diag_ventas_2200)}")
    sum_by_doc = {}
    sum_by_vend = {}
    for v in diag_ventas_2200:
        doc = v.get('docto', '')
        cant = v.get('cant', 0)
        vend = v.get('vendedor', '')
        pref = doc.split('-')[0].strip()
        sum_by_doc[pref] = sum_by_doc.get(pref, 0) + cant
        sum_by_vend[vend] = sum_by_vend.get(vend, 0) + cant
    log.info(f"  -> [DIAGNOSTICO 2200] Cantidades por tipo de documento: {sum_by_doc}")
    log.info(f"  -> [DIAGNOSTICO 2200] Cantidades por vendedor: {sum_by_vend}")

    if ventas_lineas:
        guardar_excel(path_ventas, 'IN38C_1_VtasxVendedor_LINEA_Doc', 
                      ['vendedor', 'code01', 'docto', 'fecha', 'ref', 'articulo', 'cant', 'cliente', 'valor', 'iva', 'total1', 'detalle', 'observa', 'remision', 'q_adi', 'n_adi', 'uni_factor', 'cod_benf', 'emp'], 
                      ventas_lineas)
    else:
        log.warning(f"  -> No hay registros para Ventas por Linea. Se omite la sobrescritura de {FILE_VENTAS}.")
        
    if inventario_rows:
        guardar_excel(path_inventario, 'Inventario', 
                      ['cod_item', 'referencia', 'articulo', 'linea', 'unimed', 'stock_actual', 'stock_min', 'stock_max', 'lista5', 'lista4'], 
                      inventario_rows)
    else:
        log.warning(f"  -> No hay registros para Inventario. Se omite la sobrescritura de {FILE_INVENTARIO}.")

    if cartera_documentos:
        headers_clientes = ['cod_benf', 'nit', 'nombre', 'cupo_asignado', 'saldo_total', 'saldo_corriente', 'saldo_vencido', 'cupo_disponible', 'mora_maxima', 'cod_vend', 'nombre_vend']
        headers_docs = ['cod_benf', 'nit', 'nombre', 'documento', 'fecha_elab', 'fecha_vcto', 'dias_mora', 'valor_original', 'abono', 'saldo', 'estado', 'cod_vend', 'nombre_vend']
        guardar_excel_multisheet(path_cartera, [
            ('Resumen_Clientes', headers_clientes, list(cartera_clientes.values())),
            ('Detalle_Documentos', headers_docs, cartera_documentos)
        ])
    else:
        log.warning(f"  -> No hay registros para Cartera. Se omite la sobrescritura de {FILE_CARTERA}.")

    # Sincronizar fecha en last_update.json
    now = datetime.now()
    fecha_sync = now.strftime('%d/%m/%Y %I:%M %p')
    last_update_path = os.path.join(PROYECTO_DIR, 'last_update.json')
    log.info(f"Guardando fecha de ultima sincronizacion ({fecha_sync}) en last_update.json...")
    with open(last_update_path, 'w', encoding='utf-8') as f:
        json.dump({'last_update': fecha_sync}, f)

    # =========================================================================
    # 9. Copiar archivos a public/ y dist/
    # =========================================================================
    public_dir = os.path.join(PROYECTO_DIR, 'public')
    dist_dir = os.path.join(PROYECTO_DIR, 'dist')
    
    os.makedirs(public_dir, exist_ok=True)
    os.makedirs(dist_dir, exist_ok=True)
    
    archivos_a_copiar = []
    if clientes_maestra and vendedores:
        archivos_a_copiar.append(FILE_MAESTRA)
    if ventas_lineas:
        archivos_a_copiar.append(FILE_VENTAS)
    if inventario_rows:
        archivos_a_copiar.append(FILE_INVENTARIO)
    if cartera_documentos:
        archivos_a_copiar.append(FILE_CARTERA)
    archivos_a_copiar.append('last_update.json')
    
    # Copiar a public/
    for name in archivos_a_copiar:
        src = os.path.join(PROYECTO_DIR, name)
        dst = os.path.join(public_dir, name)
        try:
            shutil.copy2(src, dst)
            log.info(f"  -> Copiado a public/: {name}")
        except Exception as e:
            log.error(f"Error al copiar {name} a public/: {e}")
            
    # Sincronizar en dist/
    log.info("Sincronizando archivos con carpeta de produccion (dist/ local)...")
    for name in archivos_a_copiar:
        src = os.path.join(PROYECTO_DIR, name)
        dst = os.path.join(dist_dir, name)
        try:
            shutil.copy2(src, dst)
            log.info(f"  -> Sincronizado en dist/ local con exito.")
        except Exception as e:
            log.error(f"Error al copiar {name} a dist/: {e}")
            
    # Sincronizar en C:\inetpub\wwwroot si existe en el servidor
    iis_dir = r'C:\inetpub\wwwroot'
    if os.path.exists(iis_dir):
        log.info(f"Sincronizando archivos con la carpeta del servidor IIS ({iis_dir})...")
        for name in archivos_a_copiar:
            src = os.path.join(PROYECTO_DIR, name)
            dst = os.path.join(iis_dir, name)
            try:
                shutil.copy2(src, dst)
                log.info(f"  -> Sincronizado en IIS ({iis_dir}) con exito: {name}")
            except Exception as e:
                log.error(f"Error al copiar {name} a IIS: {e}")
                
    log.info("=== PROCESO COMPLETADO CON EXITO ===")

if __name__ == '__main__':
    try:
        exportar_datos()
    except Exception as e:
        log.exception(f"Error fatal en el proceso de actualizacion: {e}")
