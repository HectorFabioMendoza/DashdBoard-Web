# =============================================================================
# SCRIPT DE RECONCILIACIÓN - Comparación Fina entre ERP Manual y Dashboard DBF
# Uso: python reconciliar_datos.py --manual RUTA_MANUAL --script RUTA_SCRIPT
# =============================================================================

import os
import sys
import argparse
from datetime import datetime, date
import openpyxl

# Helper para buscar campos de forma insensible a mayúsculas/minúsculas y sinónimos
def get_row_value(row, keys_to_try, headers_map, default=None):
    for k in keys_to_try:
        ku = k.lower().strip()
        if ku in headers_map:
            col_idx = headers_map[ku]
            val = row[col_idx]
            if val is not None:
                return val
    return default

def obtener_serial_fecha(d):
    if d is None:
        return 0
    if isinstance(d, datetime):
        d = d.date()
    if isinstance(d, date):
        fecha_base = date(1899, 12, 30)
        return (d - fecha_base).days
    try:
        return int(d)
    except:
        return 0

def reconciliar(manual_path, script_path, output_report_path):
    print(f"Leyendo archivo manual del ERP: {manual_path}")
    wb_m = openpyxl.load_workbook(manual_path, data_only=True)
    sheet_m = wb_m.active
    print(f"  → Hoja activa: {sheet_m.title} ({sheet_m.max_row} filas)")

    print(f"Leyendo archivo generado por el script: {script_path}")
    wb_s = openpyxl.load_workbook(script_path, data_only=True)
    sheet_s = wb_s.active
    print(f"  → Hoja activa: {sheet_s.title} ({sheet_s.max_row} filas)")

    # 1. Obtener encabezados
    rows_m = list(sheet_m.iter_rows(values_only=True))
    rows_s = list(sheet_s.iter_rows(values_only=True))

    if not rows_m or not rows_s:
        print("Error: Uno de los archivos está vacío.")
        return

    header_m = [str(cell).lower().strip() if cell is not None else "" for cell in rows_m[0]]
    header_s = [str(cell).lower().strip() if cell is not None else "" for cell in rows_s[0]]

    map_m = {name: idx for idx, name in enumerate(header_m) if name}
    map_s = {name: idx for idx, name in enumerate(header_s) if name}

    # Definir campos a buscar
    synonyms_docto = ['docto', 'documento', 'docto1', 'nro_doc', 'nro']
    synonyms_total = ['total1', 'total', 'valor', 'neto', 'total_neto']
    synonyms_cant  = ['cant', 'cantidad', 'unidades']
    synonyms_ref   = ['ref', 'referencia', 'codigo', 'cod_prod']
    synonyms_fec   = ['fecha', 'fecha_doc', 'fech']

    # 2. Cargar datos
    data_manual = {}
    data_script = {}

    # Determinar si es reporte de detalle (Ventas por línea) o de cabecera (Maestra)
    is_line_detail = any(x in map_m for x in ['ref', 'referencia', 'codigo']) or any(x in map_s for x in ['ref', 'referencia', 'codigo'])

    if is_line_detail:
        print("Detectado formato: DETALLE DE VENTAS POR LÍNEA")
        # Agrupar por clave compuesta: docto + ref + cliente/nit
        synonyms_benf = ['cod_benf', 'cliente', 'nit', 'cod_cli', 'cod_client']

        # Cargar manual
        for i, row in enumerate(rows_m[1:], start=2):
            doc = get_row_value(row, synonyms_docto, map_m)
            ref = get_row_value(row, synonyms_ref, map_m)
            benf = get_row_value(row, synonyms_benf, map_m, '')
            val = get_row_value(row, synonyms_total, map_m, 0.0)
            cant = get_row_value(row, synonyms_cant, map_m, 0.0)
            
            if not doc or not ref:
                continue

            doc_key = str(doc).strip().upper()
            ref_key = str(ref).strip().upper()
            benf_key = str(benf).strip().upper()
            key = (doc_key, ref_key, benf_key)

            if key not in data_manual:
                data_manual[key] = {'val': 0.0, 'cant': 0.0, 'filas': []}
            data_manual[key]['val'] += float(val or 0)
            data_manual[key]['cant'] += float(cant or 0)
            data_manual[key]['filas'].append(i)

        # Cargar script
        for i, row in enumerate(rows_s[1:], start=2):
            doc = get_row_value(row, synonyms_docto, map_s)
            ref = get_row_value(row, synonyms_ref, map_s)
            benf = get_row_value(row, synonyms_benf, map_s, '')
            val = get_row_value(row, synonyms_total, map_s, 0.0)
            cant = get_row_value(row, synonyms_cant, map_s, 0.0)
            
            if not doc or not ref:
                continue

            doc_key = str(doc).strip().upper()
            ref_key = str(ref).strip().upper()
            benf_key = str(benf).strip().upper()
            key = (doc_key, ref_key, benf_key)

            if key not in data_script:
                data_script[key] = {'val': 0.0, 'cant': 0.0, 'filas': []}
            data_script[key]['val'] += float(val or 0)
            data_script[key]['cant'] += float(cant or 0)
            data_script[key]['filas'].append(i)

    else:
        print("Detectado formato: CABECERA / MAESTRA DE CLIENTES")
        # Agrupar por clave sencilla: docto
        for i, row in enumerate(rows_m[1:], start=2):
            doc = get_row_value(row, synonyms_docto, map_m)
            val = get_row_value(row, synonyms_total, map_m, 0.0)
            if not doc:
                continue
            key = str(doc).strip().upper()
            if key not in data_manual:
                data_manual[key] = {'val': 0.0, 'filas': []}
            data_manual[key]['val'] += float(val or 0)
            data_manual[key]['filas'].append(i)

        for i, row in enumerate(rows_s[1:], start=2):
            doc = get_row_value(row, synonyms_docto, map_s)
            val = get_row_value(row, synonyms_total, map_s, 0.0)
            if not doc:
                continue
            key = str(doc).strip().upper()
            if key not in data_script:
                data_script[key] = {'val': 0.0, 'filas': []}
            data_script[key]['val'] += float(val or 0)
            data_script[key]['filas'].append(i)

    # 3. Conciliar
    keys_manual = set(data_manual.keys())
    keys_script = set(data_script.keys())

    only_in_manual = keys_manual - keys_script
    only_in_script = keys_script - keys_manual
    common_keys = keys_manual & keys_script

    discrepancies = []
    total_val_manual = 0.0
    total_val_script = 0.0

    # Sumatorias globales
    for key, data in data_manual.items():
        total_val_manual += data['val']
    for key, data in data_script.items():
        total_val_script += data['val']

    # Verificar diferencias de valor en registros comunes
    for key in common_keys:
        val_m = data_manual[key]['val']
        val_s = data_script[key]['val']
        diff = val_m - val_s

        if is_line_detail:
            cant_m = data_manual[key]['cant']
            cant_s = data_script[key]['cant']
            diff_cant = cant_m - cant_s
            if abs(diff) > 1.0 or abs(diff_cant) > 0.01:
                discrepancies.append({
                    'key': key,
                    'manual_val': val_m,
                    'script_val': val_s,
                    'diff_val': diff,
                    'manual_cant': cant_m,
                    'script_cant': cant_s,
                    'diff_cant': diff_cant
                })
        else:
            if abs(diff) > 1.0: # Margen de $1 peso por redondeos
                discrepancies.append({
                    'key': key,
                    'manual_val': val_m,
                    'script_val': val_s,
                    'diff_val': diff
                })

    # 4. Generar reporte
    with open(output_report_path, 'w', encoding='utf-8') as f:
        f.write("=============================================================================\n")
        f.write("REPORTE DE CONCILIACIÓN DE DATOS (ERP MANUAL VS DASHBOARD DBF)\n")
        f.write(f"Fecha de reporte: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=============================================================================\n\n")

        f.write("--- RESUMEN GLOBAL ---\n")
        f.write(f"Suma Total Manual ERP   : ${total_val_manual:,.2f}\n")
        f.write(f"Suma Total Script DBF   : ${total_val_script:,.2f}\n")
        f.write(f"Diferencia (Manual-Scr) : ${total_val_manual - total_val_script:,.2f}\n")
        f.write(f"Registros en ERP Manual : {len(keys_manual)}\n")
        f.write(f"Registros en Script DBF : {len(keys_script)}\n\n")

        f.write(f"--- ANÁLISIS DE REGISTROS ---\n")
        f.write(f"1. Registros en ERP Manual pero FALTANTES en Script: {len(only_in_manual)}\n")
        f.write(f"2. Registros en Script pero FALTANTES en ERP Manual: {len(only_in_script)}\n")
        f.write(f"3. Registros con diferencias de VALOR/CANTIDAD: {len(discrepancies)}\n\n")

        if only_in_manual:
            f.write("-----------------------------------------------------------------------------\n")
            f.write("DETALLE: REGISTROS EN ERP MANUAL FALTANTES EN EL SCRIPT\n")
            f.write("-----------------------------------------------------------------------------\n")
            for k in list(only_in_manual)[:50]:
                if is_line_detail:
                    f.write(f"Clave: Doc={k[0]}, Ref={k[1]}, Cliente={k[2]} | Valor ERP: ${data_manual[k]['val']:,.2f} | Cant ERP: {data_manual[k]['cant']} | Filas origen: {data_manual[k]['filas']}\n")
                else:
                    f.write(f"Doc: {k} | Valor ERP: ${data_manual[k]['val']:,.2f} | Filas origen: {data_manual[k]['filas']}\n")
            if len(only_in_manual) > 50:
                f.write(f"... y {len(only_in_manual) - 50} registros más.\n")
            f.write("\n")

        if only_in_script:
            f.write("-----------------------------------------------------------------------------\n")
            f.write("DETALLE: REGISTROS EN EL SCRIPT FALTANTES EN EL ERP MANUAL\n")
            f.write("-----------------------------------------------------------------------------\n")
            for k in list(only_in_script)[:50]:
                if is_line_detail:
                    f.write(f"Clave: Doc={k[0]}, Ref={k[1]}, Cliente={k[2]} | Valor Script: ${data_script[k]['val']:,.2f} | Cant Script: {data_script[k]['cant']} | Filas origen: {data_script[k]['filas']}\n")
                else:
                    f.write(f"Doc: {k} | Valor Script: ${data_script[k]['val']:,.2f} | Filas origen: {data_script[k]['filas']}\n")
            if len(only_in_script) > 50:
                f.write(f"... y {len(only_in_script) - 50} registros más.\n")
            f.write("\n")

        if discrepancies:
            f.write("-----------------------------------------------------------------------------\n")
            f.write("DETALLE: REGISTROS CON VALORES / CANTIDADES CON DIFERENCIA\n")
            f.write("-----------------------------------------------------------------------------\n")
            for d in discrepancies[:50]:
                k = d['key']
                if is_line_detail:
                    f.write(f"Clave: Doc={k[0]}, Ref={k[1]}, Cliente={k[2]}\n")
                    f.write(f"  → ERP Manual: Valor: ${d['manual_val']:,.2f}, Cant: {d['manual_cant']} | Filas: {data_manual[k]['filas']}\n")
                    f.write(f"  → Script DBF: Valor: ${d['script_val']:,.2f}, Cant: {d['script_cant']} | Filas: {data_script[k]['filas']}\n")
                    f.write(f"  → Desviación: Valor: ${d['diff_val']:,.2f}, Cant: {d['diff_cant']}\n")
                else:
                    f.write(f"Doc: {k}\n")
                    f.write(f"  → ERP Manual: ${d['manual_val']:,.2f} | Filas: {data_manual[k]['filas']}\n")
                    f.write(f"  → Script DBF: ${d['script_val']:,.2f} | Filas: {data_script[k]['filas']}\n")
                    f.write(f"  → Desviación: ${d['diff_val']:,.2f}\n")
            if len(discrepancies) > 50:
                f.write(f"... y {len(discrepancies) - 50} registros más.\n")
            f.write("\n")

    print("\n====================================================")
    print("CONCILIACIÓN COMPLETADA")
    print(f"Total ERP Manual   : ${total_val_manual:,.2f}")
    print(f"Total Script DBF   : ${total_val_script:,.2f}")
    print(f"Diferencia neta    : ${total_val_manual - total_val_script:,.2f}")
    print(f"Reporte detallado guardado en: {output_report_path}")
    print("====================================================")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Reconciliación de datos de ventas.")
    parser.add_argument('--manual', required=True, help="Ruta al archivo Excel manual descargado del ERP.")
    parser.add_argument('--script', required=True, help="Ruta al archivo Excel generado por el script.")
    parser.add_argument('--out', default="reporte_conciliacion.txt", help="Ruta del archivo de texto de salida para el reporte.")

    args = parser.parse_args()

    if not os.path.exists(args.manual):
        print(f"Error: No existe el archivo manual en {args.manual}")
        sys.exit(1)
    if not os.path.exists(args.script):
        print(f"Error: No existe el archivo del script en {args.script}")
        sys.exit(1)

    reconciliar(args.manual, args.script, args.out)
