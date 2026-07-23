# =============================================================================
# SCRIPT DE ENVIO - Alertas de Inventario y Abastecimiento por correo
# Lee los mismos Inventario.xlsx / Ventas por linea.xlsx que genera
# actualizar_dashboard_dbf.py y envia a gerencia un resumen HTML con los
# productos de mayor urgencia de compra (sin costos de Lista 4 / Lista 5).
#
# Se llama desde actualizar_datos_dbf_silencioso.bat en las 4 corridas diarias
# (7am/1pm/5pm/10pm) programadas por programar_actualizacion_dbf.ps1, pero solo
# envia realmente el correo si se ejecuta dentro de la franja de las 7am (para
# no repetir el aviso en cada corrida). Usa --forzar para enviar manualmente
# sin importar la hora (pruebas).
#
# Uso: python enviar_alertas_inventario.py [--forzar]
# =============================================================================

import os
import sys
import json
import time
import socket
import argparse
import logging
import smtplib
import calendar
from datetime import datetime, timedelta
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr

try:
    import openpyxl
except ImportError:
    print("Instalando libreria openpyxl...")
    os.system('pip install openpyxl')
    import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, 'config_alertas_inventario.json')
CONFIG_LOCAL_FILE = os.path.join(SCRIPT_DIR, 'config_alertas_inventario.local.json')
LOG_FILE = os.path.join(SCRIPT_DIR, 'log_alertas_inventario.txt')

FILE_INVENTARIO = 'Inventario.xlsx'
FILE_VENTAS = 'Ventas por linea.xlsx'
FILE_MAESTRA = '1Maestra de clientes2026.xlsx'

TEAL = '#0f5c52'

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
log = logging.getLogger('alertas_inventario')


def clean_vendor_name_local(v):
    if not v:
        return ''
    import re
    return re.sub(r'^\d+\s+', '', str(v)).strip().upper()


def cargar_vendedores_validos(ruta_datos):
    """Extrae el conjunto de nombres de asesores/vendedores validos desde la
    Maestra de Clientes, replicando exactamente selectedVendors en App.tsx."""
    ruta_maestra = os.path.join(ruta_datos, FILE_MAESTRA)
    if not os.path.exists(ruta_maestra):
        return None
    try:
        wb = openpyxl.load_workbook(ruta_maestra, read_only=True)
        ws = wb[wb.sheetnames[0]]
        filas = ws.iter_rows(values_only=True)
        encabezado = next(filas)
        idx = {nombre: i for i, nombre in enumerate(encabezado)}
        idx_vend = idx.get('vendedor')
        if idx_vend is None:
            wb.close()
            return None

        vendedores = set()
        for fila in filas:
            v_raw = fila[idx_vend]
            if v_raw:
                v_clean = clean_vendor_name_local(v_raw)
                if v_clean:
                    vendedores.add(v_clean)
        wb.close()
        return vendedores
    except Exception as e:
        log.warning(f"No se pudo cargar vendedores desde {FILE_MAESTRA}: {e}")
        return None


def _config_por_defecto():
    return {
        "top_n": 25,
        "email": {
            "servidor_smtp": "smtp.gmail.com",
            "puerto_smtp": 587,
            "remitente": "tu_usuario@gmail.com",
            "password": "tu_password_o_token",
            "nombre_remitente": "Alertas de Inventario y Abastecimiento",
            "destinatario": "hectorfabio.mendoza@gmail.com",
            "cc": ""
        }
    }


def cargar_configuracion():
    """Carga config_alertas_inventario.json (o una por defecto) y superpone las
    credenciales de correo desde config_alertas_inventario.local.json si existe.
    Ese archivo local nunca se versiona (ver .gitignore) para no exponer la
    contrasena SMTP en el repositorio."""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
        except Exception as e:
            log.error(f"Error al leer config_alertas_inventario.json: {e}")
            config = _config_por_defecto()
    else:
        config = _config_por_defecto()

    if os.path.exists(CONFIG_LOCAL_FILE):
        try:
            with open(CONFIG_LOCAL_FILE, 'r', encoding='utf-8') as f:
                config_local = json.load(f)
            config.setdefault("email", {}).update(config_local.get("email", {}))
        except Exception as e:
            log.error(f"Error al leer config_alertas_inventario.local.json: {e}")
    else:
        log.warning(
            "No se encontro config_alertas_inventario.local.json (credenciales SMTP). "
            "Copia config_alertas_inventario.json a config_alertas_inventario.local.json "
            "y completa remitente/password con las credenciales reales."
        )

    return config


def resolver_ruta_datos():
    """Los .xlsx que consume el frontend viven en la raiz del proyecto (junto a
    este script) en el servidor; en desarrollo local caen en public/."""
    candidatos = [SCRIPT_DIR, os.path.join(SCRIPT_DIR, 'public')]
    for base in candidatos:
        if os.path.exists(os.path.join(base, FILE_INVENTARIO)):
            return base
    raise FileNotFoundError(
        f"No se encontro {FILE_INVENTARIO} en {candidatos}. "
        "Corre primero actualizar_dashboard_dbf.py."
    )


def excel_a_fecha(serial):
    return datetime(1899, 12, 30) + timedelta(days=int(serial))


def cargar_velocidad_ventas(ruta_datos):
    """Replica la logica de productSalesVelocityMap en App.tsx, incluyendo el
    filtro por asesores comerciales (selectedVendors en App.tsx) y el
    prefiltro automatico de 'Periodos' en una sola pasada acelerada de lectura."""
    ruta_ventas = os.path.join(ruta_datos, FILE_VENTAS)
    vendedores_validos = cargar_vendedores_validos(ruta_datos)

    wb = openpyxl.load_workbook(ruta_ventas, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    encabezado = next(filas)
    idx = {nombre: i for i, nombre in enumerate(encabezado)}

    idx_vendedor = idx.get('vendedor')
    idx_fecha = idx.get('fecha')
    idx_ref = idx.get('ref')
    idx_cant = idx.get('cant')

    registros = []
    max_serial = 0
    meses_vistos = set()

    for fila in filas:
        if vendedores_validos and idx_vendedor is not None:
            v_clean = clean_vendor_name_local(fila[idx_vendedor])
            if v_clean not in vendedores_validos:
                continue

        fecha = fila[idx_fecha] if idx_fecha is not None else None
        if not fecha:
            continue

        try:
            d = excel_a_fecha(fecha)
            mes_tuple = (d.year, d.month)
            if fecha > max_serial:
                max_serial = fecha
            meses_vistos.add(mes_tuple)

            ref = str(fila[idx_ref] or '').strip().lstrip('0') if idx_ref is not None else ''
            cant = float(fila[idx_cant] or 0) if idx_cant is not None else 0.0
            if ref and cant != 0:
                registros.append((mes_tuple, ref, cant))
        except Exception:
            pass
    wb.close()

    if not registros:
        return {}

    mes_actual = None
    proporcion_mes_actual = 1.0
    if max_serial > 0:
        fecha_max = excel_a_fecha(max_serial)
        mes_actual = (fecha_max.year, fecha_max.month)
        dias_en_mes = calendar.monthrange(fecha_max.year, fecha_max.month)[1]
        proporcion_mes_actual = max(1, fecha_max.day) / dias_en_mes

    hoy = datetime.now()
    mes_hoy = (hoy.year, hoy.month)
    n_meses_ventana = 4 if mes_hoy in meses_vistos else 3
    meses_seleccionados = set(sorted(meses_vistos)[-n_meses_ventana:])

    cantidad_por_ref = defaultdict(float)
    for mes_tuple, ref, cant in registros:
        if mes_tuple in meses_seleccionados:
            cantidad_por_ref[ref] += cant

    divisor = 0.0
    for mes in meses_seleccionados:
        divisor += proporcion_mes_actual if mes == mes_actual else 1.0
    if divisor < 0.05:
        divisor = 1.0

    return {ref: cantidad / divisor for ref, cantidad in cantidad_por_ref.items()}


def cargar_inventario_priorizado(ruta_datos):
    """Lee Inventario.xlsx y calcula velocidad, prioridad, estado y cobertura
    para cada articulo, con la misma logica de negocio de App.tsx (seccion
    'processedInventory' e 'inventoryKPIs')."""
    velocidad_por_ref = cargar_velocidad_ventas(ruta_datos)

    wb = openpyxl.load_workbook(os.path.join(ruta_datos, FILE_INVENTARIO), read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    encabezado = next(filas)
    idx = {nombre: i for i, nombre in enumerate(encabezado)}

    items = []
    for fila in filas:
        cod_raw = str(fila[idx['cod_item']] or '').strip()
        ref_raw = str(fila[idx['referencia']] or '').strip()
        cod_clean = cod_raw.lstrip('0')
        ref_clean = ref_raw.lstrip('0')
        articulo = str(fila[idx['articulo']] or '').strip()
        linea = str(fila[idx['linea']] or '').strip()
        stock = float(fila[idx['stock_actual']] or 0)

        # Replicar exactamente la resolucion de velocidad de App.tsx:
        # productSalesVelocityMap[cleanCod] || productSalesVelocityMap[cleanRef] || 0
        velocidad = velocidad_por_ref.get(cod_clean) or velocidad_por_ref.get(ref_clean) or 0.0
        prioridad = velocidad / (stock + 1)
        ventas_diarias = velocidad / 30
        cobertura_dias = stock / ventas_diarias if ventas_diarias > 0 else 999

        if stock <= 0:
            estado = 'Agotado'
        elif velocidad > 0 and cobertura_dias <= 15:
            estado = 'Riesgo'
        elif velocidad > 0 and cobertura_dias <= 30:
            estado = 'Atencion'
        else:
            estado = 'Saludable'

        items.append({
            'referencia': ref_raw or cod_raw,
            'articulo': articulo,
            'linea': linea,
            'stock': stock,
            'velocidad': velocidad,
            'prioridad': prioridad,
            'cobertura_dias': cobertura_dias,
            'estado': estado,
        })

    return items


def calcular_kpis(items):
    """Solo cuenta como alerta los articulos con ventas registradas
    (velocidad > 0), igual que inventoryKPIs en App.tsx."""
    agotados = riesgo = atencion = 0
    suma_cobertura = 0.0
    items_con_ventas = 0

    for it in items:
        if it['velocidad'] > 0:
            if it['estado'] == 'Agotado':
                agotados += 1
            elif it['estado'] == 'Riesgo':
                riesgo += 1
            elif it['estado'] == 'Atencion':
                atencion += 1

            items_con_ventas += 1
            suma_cobertura += min(365, it['cobertura_dias'])

    cobertura_promedio = round(suma_cobertura / items_con_ventas) if items_con_ventas else 0
    return {
        'agotados': agotados,
        'riesgo': riesgo,
        'atencion': atencion,
        'cobertura_promedio': cobertura_promedio,
    }


def fmt_entero_es(n):
    return f"{round(n):,}".replace(",", ".")


def fmt_decimal_es(n):
    return f"{n:,.1f}".replace(",", "@").replace(".", ",").replace("@", ".")


ESTADO_ESTILO = {
    'Agotado': ('#fee2e2', '#b91c1c', 'AGOTADO'),
    'Riesgo': ('#ffedd5', '#c2410c', 'RIESGO'),
    'Atencion': ('#fef9c3', '#a16207', 'ATENCIÓN'),
}


def generar_fila_html(item, ranking):
    bg, color, etiqueta = ESTADO_ESTILO.get(item['estado'], ('#eef4fa', '#334155', item['estado'].upper()))
    linea = item['linea'][5:].strip().title() if len(item['linea']) > 5 and item['linea'][:4].isdigit() else item['linea'].title()
    if not linea or linea.lower() == 'linea general':
        linea = 'Línea General'
    return f"""
              <tr>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;">{item['referencia']}</td>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;">{item['articulo'].title()}</td>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;color:#64748b;">{linea}</td>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;text-align:center;"><span style="background:{bg};color:{color};border-radius:4px;padding:1px 6px;font-size:10px;font-weight:700;">{etiqueta}</span></td>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;text-align:right;">{fmt_decimal_es(item['velocidad'])}</td>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;text-align:right;">{fmt_entero_es(item['stock'])}</td>
                <td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;text-align:right;font-weight:700;">#{ranking}</td>
              </tr>"""


def generar_tarjeta_kpi_html(etiqueta, valor, bg, color, ultima=False):
    pad_right = "" if ultima else "padding-right:8px;"
    return f"""
              <td width="25%" style="{pad_right}" valign="top">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:{bg};border-radius:8px;">
                  <tr><td style="padding:7px 10px;">
                    <div style="font-size:9px;color:{color};letter-spacing:.3px;text-transform:uppercase;">{etiqueta}</div>
                    <div style="font-size:18px;font-weight:700;color:{color};margin-top:2px;">{valor}</div>
                  </td></tr>
                </table>
              </td>"""


def generar_cuerpo_html(items, kpis, top_n, fecha_larga, total_alertas):
    top_items = sorted(
        (it for it in items if it['velocidad'] > 0),
        key=lambda it: it['prioridad'],
        reverse=True,
    )[:top_n]

    filas_html = ''.join(generar_fila_html(it, i + 1) for i, it in enumerate(top_items))

    return f"""\
<html>
  <body style="font-family:'Segoe UI',Arial,sans-serif;color:#1f2937;margin:0;padding:0;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:{TEAL};border-radius:6px 6px 0 0;">
      <tr>
        <td style="padding:14px 20px;">
          <div style="color:#ffffff;font-size:17px;font-weight:700;">Alertas de Inventario y Abastecimiento</div>
          <div style="color:#bfe3dc;font-size:11.5px;margin-top:2px;">Distribuidora JR &nbsp;&middot;&nbsp; Corte: {fecha_larga}</div>
        </td>
      </tr>
    </table>

    <div style="padding:16px 20px;background:#ffffff;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin:0 0 14px;">
        <tr>
          {generar_tarjeta_kpi_html("Agotados", fmt_entero_es(kpis['agotados']), '#fee2e2', '#b91c1c')}
          {generar_tarjeta_kpi_html("Crítico &lt;15d", fmt_entero_es(kpis['riesgo']), '#ffedd5', '#c2410c')}
          {generar_tarjeta_kpi_html("Atención &lt;30d", fmt_entero_es(kpis['atencion']), '#fef9c3', '#a16207')}
          {generar_tarjeta_kpi_html("Cobert. Prom.", f"{kpis['cobertura_promedio']} días", '#eef4fa', TEAL, ultima=True)}
        </tr>
      </table>

      <h3 style="color:{TEAL};font-size:13.5px;margin:0 0 3px;">Top {len(top_items)} &mdash; Mayor urgencia de compra</h3>
      <div style="font-size:10.5px;color:#94a3b8;margin:0 0 6px;">Prioridad de compra sobre los {fmt_entero_es(total_alertas)} ítems en alerta (#1 = más urgente), según Ventas Prom. Mes &divide; (Stock Actual + 1)</div>
      <table style="border-collapse:collapse;width:100%;font-size:11.5px;line-height:1.25;">
        <thead>
          <tr style="background:{TEAL};color:#ffffff;">
            <th style="padding:5px 6px;text-align:left;">Ref.</th>
            <th style="padding:5px 6px;text-align:left;">Producto</th>
            <th style="padding:5px 6px;text-align:left;">Línea / Marca</th>
            <th style="padding:5px 6px;text-align:center;">Estado</th>
            <th style="padding:5px 6px;text-align:right;">Ventas Prom./Mes</th>
            <th style="padding:5px 6px;text-align:right;">Stock</th>
            <th style="padding:5px 6px;text-align:right;">Prioridad</th>
          </tr>
        </thead>
        <tbody>{filas_html}
        </tbody>
      </table>

      <table role="presentation" cellpadding="0" cellspacing="0" style="margin:10px 0 0;">
        <tr>
          <td style="background:#eef2f7;border:1px solid #dbe4ee;border-radius:6px;padding:5px 10px;font-size:11px;color:#334155;">
            Este es un extracto: hay <strong>{fmt_entero_es(total_alertas)} ítems</strong> en alerta en total dentro del Dashboard.
          </td>
        </tr>
      </table>

      <p style="margin:14px 0 2px;color:#94a3b8;font-size:11px;">Hola, este es un resumen automático de los productos con mayor urgencia de reabastecimiento según el Dashboard de Inventario. Ingresa a la herramienta para iniciar el proceso de compra con los proveedores correspondientes. Este es un correo automático generado por el sistema.</p>
    </div>
  </body>
</html>
"""


MESES_ES = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
            'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']


def fmt_fecha_larga_es(d):
    return f"{d.day} de {MESES_ES[d.month - 1]} de {d.year}"


def enviar_email(cuerpo_html, config):
    email_cfg = config.get("email", {})
    smtp_server = str(email_cfg.get("servidor_smtp", "smtp.gmail.com")).strip()
    smtp_port = email_cfg.get("puerto_smtp", 587)
    sender_email = str(email_cfg.get("remitente", "")).strip()
    sender_password = str(email_cfg.get("password", "")).strip()
    sender_name = str(email_cfg.get("nombre_remitente", "Alertas de Inventario y Abastecimiento")).strip()
    recipient_email = str(email_cfg.get("destinatario", "hectorfabio.mendoza@gmail.com")).strip()
    cc_email = str(email_cfg.get("cc", "")).strip()

    if "tu_usuario@gmail.com" in sender_email or not sender_email or not sender_password:
        log.warning("Servidor SMTP o credenciales no configurados en config_alertas_inventario.local.json.")
        log.warning("Saltando envío.")
        return False

    dest_str = recipient_email
    if cc_email:
        dest_str += f" con copia a {cc_email}"
    log.info(f"Iniciando envío de correo a {dest_str} vía {smtp_server}:{smtp_port}...")

    try:
        mensaje = MIMEMultipart('alternative')
        mensaje['From'] = formataddr((sender_name, sender_email))
        mensaje['To'] = recipient_email
        if cc_email:
            mensaje['Cc'] = cc_email

        to_addrs = [r.strip() for r in recipient_email.split(',') if r.strip()]
        if cc_email:
            to_addrs.extend([r.strip() for r in cc_email.split(',') if r.strip()])

        fecha_str = datetime.now().strftime("%Y-%m-%d")
        mensaje['Subject'] = f"Alertas de Inventario y Abastecimiento - Distribuidora JR - {fecha_str}"

        cuerpo_plain = (
            "Hola,\n\n"
            "Este es el resumen automático de los productos con mayor urgencia de "
            "reabastecimiento del Dashboard de Inventario de Distribuidora JR.\n\n"
            "Ingresa al Dashboard para ver el listado completo y gestionar la compra "
            "con los proveedores correspondientes.\n"
        )
        mensaje.attach(MIMEText(cuerpo_plain, 'plain', 'utf-8'))
        mensaje.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))

        # Conexión SMTP con reintentos y fallback a IPs directas de Gmail ante fallos de DNS (Errno 11001)
        ips_gmail_fallback = ['74.125.26.109', '142.250.115.108', '173.194.76.108']
        max_intentos = 3
        for intento in range(1, max_intentos + 1):
            try:
                try:
                    server = smtplib.SMTP(smtp_server, smtp_port, timeout=15)
                except (socket.gaierror, socket.error, TimeoutError) as net_err:
                    log.warning(f"Conexión a '{smtp_server}' por dominio falló ({net_err}). Intentando resolución por IPv4...")
                    try:
                        ip_resuelta = socket.gethostbyname(smtp_server)
                        server = smtplib.SMTP(ip_resuelta, smtp_port, timeout=15)
                    except Exception:
                        log.warning("Servicio DNS local no disponible. Conectando a IP directa de servidor SMTP...")
                        server = None
                        for ip in ips_gmail_fallback:
                            try:
                                server = smtplib.SMTP(ip, smtp_port, timeout=10)
                                break
                            except Exception:
                                continue
                        if not server:
                            raise net_err

                server.starttls()
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, to_addrs, mensaje.as_string())
                server.quit()
                log.info(f"¡Correo enviado exitosamente a {recipient_email}!")
                return True
            except Exception as err_smtp:
                if intento < max_intentos:
                    log.warning(f"Intento {intento}/{max_intentos} falló al conectar a SMTP ({err_smtp}). Reintentando en 3s...")
                    time.sleep(3)
                else:
                    raise err_smtp
    except Exception as e:
        log.error(f"Error al enviar el correo electrónico: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description="Envio de Alertas de Inventario y Abastecimiento.")
    parser.add_argument(
        '--forzar', action='store_true',
        help="Envia el correo sin importar la hora actual (uso manual/pruebas)."
    )
    args = parser.parse_args()

    hora_actual = datetime.now().hour
    if not args.forzar and hora_actual != 7:
        log.info(
            f"Hora actual ({hora_actual}h) fuera de la franja de las 7am; "
            "esta corrida del Programador de Tareas no envia el correo. "
            "Usa --forzar para enviarlo de todas formas."
        )
        return

    config = cargar_configuracion()
    ruta_datos = resolver_ruta_datos()
    log.info(f"Leyendo datos de inventario desde: {ruta_datos}")

    items = cargar_inventario_priorizado(ruta_datos)
    kpis = calcular_kpis(items)
    # inventoryKPIs.totalAlerts en App.tsx solo suma Agotado + Riesgo (Atencion
    # se muestra aparte pero no cuenta como "alerta" para este total).
    total_alertas = kpis['agotados'] + kpis['riesgo']
    top_n = config.get('top_n', 25)

    log.info(
        f"KPIs: Agotados={kpis['agotados']} Crítico={kpis['riesgo']} "
        f"Atención={kpis['atencion']} Cobertura Prom={kpis['cobertura_promedio']}días "
        f"Total alertas={total_alertas}"
    )

    fecha_larga = fmt_fecha_larga_es(datetime.now().date())
    cuerpo_html = generar_cuerpo_html(items, kpis, top_n, fecha_larga, total_alertas)

    enviar_email(cuerpo_html, config)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log.exception(f"Error fatal en el proceso de envío de alertas: {e}")
